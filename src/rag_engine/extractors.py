from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from rag_engine.models import DocumentRecord


SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".html", ".htm", ".md", ".txt", ".pptx"}


def extract_document(path: Path) -> DocumentRecord:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file type: {suffix}")

    if suffix == ".pdf":
        content = _extract_pdf(path)
    elif suffix == ".docx":
        content = _extract_docx(path)
    elif suffix == ".xlsx":
        content = _extract_xlsx(path)
    elif suffix == ".pptx":
        content = _extract_pptx(path)
    elif suffix in {".html", ".htm"}:
        content = _extract_html(path)
    else:
        content = path.read_text(encoding="utf-8")

    return DocumentRecord(
        source_path=path,
        content=content,
        title=path.stem,
        file_type=suffix.lstrip("."),
    )


def discover_documents(root: Path) -> list[Path]:
    paths = [path for path in root.rglob("*") if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS]
    return sorted(paths)


def _extract_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(pages)


def _extract_docx(path: Path) -> str:
    from docx import Document as WordDocument

    document = WordDocument(str(path))
    return "\n".join(paragraph.text for paragraph in document.paragraphs)


def _extract_xlsx(path: Path) -> str:
    workbook = load_workbook(filename=str(path), data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"[Sheet] {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines)


def _extract_html(path: Path) -> str:
    html = path.read_text(encoding="utf-8")
    soup = BeautifulSoup(html, "html.parser")
    return soup.get_text("\n")


def _extract_pptx(path: Path) -> str:
    from zipfile import ZipFile
    import xml.etree.ElementTree as ET

    slides: list[str] = []
    namespace = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    with ZipFile(path) as archive:
        for member in sorted(name for name in archive.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml")):
            root = ET.fromstring(archive.read(member))
            texts = [node.text for node in root.findall(".//a:t", namespace) if node.text]
            if texts:
                slides.append("\n".join(texts))
    return "\n\n".join(slides)
